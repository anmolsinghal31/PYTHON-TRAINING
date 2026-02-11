from openpyxl import load_workbook

wb = load_workbook('sales_data.xlsx')
ws = wb['2025']

# Add the header to the 4th column
ws.cell(row=2, column=4).value = "Total"

# Start from row 3 to skip the two header rows
for row in range(3, ws.max_row + 1):
    quantity = ws.cell(row=row, column=2).value
    price = ws.cell(row=row, column=3).value

    # Only multiply if both cells have values to avoid errors
    if quantity is not None and price is not None:
        ws.cell(row=row, column=4).value = quantity * price

wb.save('sales_summary_bonus.xlsx')